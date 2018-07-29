from openpyxl import Workbook
from openpyxl.styles import PatternFill

class ExcelExporter:
	""" Class that exports the participants' responses to an excel spreadsheet. """

	def __init__(self, participants, quiz_prefixes, scale_prefixes):

		self.participants = participants
		self.quiz_prefixes = quiz_prefixes
		self.scale_prefixes = scale_prefixes

		self.included_columns = []
		self.output_workbook = Workbook()


	def create_all_responses_worksheet(self):
		""" Creates a worksheet that has all of the responses to the survey in it """
		all_responses_ws = self.output_workbook.create_sheet(title = 'All Responses')

		correct_fill = PatternFill('solid', fgColor = 'ccf2b5')
		incorrect_fill = PatternFill('solid', fgColor = 'fdd3cb')

		# Write to the cells for header row
		for (index, column_name) in enumerate(self.included_columns):
			column_index = index + 1
			_ = all_responses_ws.cell(column = column_index, row = 1, value = column_name)
	
		# Write all of the participants' answers (whether they are correct or incorrect) to the worksheet
		for (index, participant) in enumerate(self.participants):

			# Have to add 2, because the row header takes up the first row in the worksheet
			row_index = index + 2

			# Iterate through all of the columns that need to be included, and find each participant's responses to these
			for (col_index, included_column) in enumerate(self.included_columns):

				# Have to add 1, because the columns in the worksheet start from 1
				column_index = col_index + 1
				responses = participant.participant_responses

				# For the name column, simply show the participant's name
				if included_column == 'name':
					_ = all_responses_ws.cell(column = column_index, row = row_index, value = participant.name)
					column_index += 1
				
				# For the stream column, just show the stream name
				elif included_column == 'stream':
					_ = all_responses_ws.cell(column = column_index, row = row_index, value = participant.stream)
					column_index += 1

				elif 'PlanToTeachIn' in included_column or 'PlanToIntegrateIn' in included_column:
					integrate_response = [r for r in responses if r.question_number == included_column]
					if (len(integrate_response) == 1):
						integrate_answer = integrate_response[0].response_answer
						_ = all_responses_ws.cell(column = column_index, row = row_index, value = integrate_answer)
						column_index += 1

					elif (len(integrate_response) == 0):
						print('No response found for the Integration Q: {} for participant {}'.format(included_column, participant.name))

					else:
						print('More than one response found for the Integration Q: {} for participant {}'.format(included_column, participant.name))

				# For the quiz question columns, we have to find the response from the participant
				else:

					# Find the response from the participant
					included_column_split = included_column.split('_')
					category = included_column_split[0]
					question_number = included_column_split[1]
					response_values = [r for r in responses if r.category == category and r.question_number == question_number]

					# Check that a response from the participant has been found
					if len(response_values) == 1:
						response = response_values[0]

						# The value printed to the spreadsehet will be 'Correct' or 'Incorrect' for quiz questions and the value for others
						if response.is_quiz_question:
							cell_value = 'Correct' if response.is_correct else 'Incorrect'
							cell_fill = correct_fill if response.is_correct else incorrect_fill
							response_cell = all_responses_ws.cell(column = column_index, row = row_index, value = cell_value)
							response_cell.fill = cell_fill
							column_index += 1
						else:
							_ = all_responses_ws.cell(column = column_index, row = row_index, value = int(response.response_answer))
							column_index += 1

					elif len(response_values) == 0:
						print('No response found for the {} Q{} for {}'.format(category, question_number, participant.name))

					else:
						print('More than one response found for the {} Q{} for {}'.format(category, question_number, participant.name))



	def create_summary_worksheet(self):
		""" Creates a worksheet that lists all of the participants and summarises their results. """

		summary_ws = self.output_workbook.active
		summary_ws.title = 'Summary'

		summary_column_names = []
		summary_column_names.append('Name')
		summary_column_names.append('Stream')

		for quiz_prefix in self.quiz_prefixes:
			summary_column_names.append('{} Total'.format(quiz_prefix.title()))

		for scale_prefix in self.scale_prefixes:
			summary_column_names.append('{} Average'.format(scale_prefix))

		# Write the column names 
		for (index, column_name) in enumerate(summary_column_names):
			column_index = index + 1
			_ = summary_ws.cell(column = column_index, row = 1, value = column_name)

		# For each participant, we will write out a summary of their quiz results and reporting of scales to a row
		for (index, participant) in enumerate(self.participants):

			row_index = index + 2
			column_index = 1

			_ = summary_ws.cell(column = column_index, row = row_index, value = participant.name)
			column_index += 1

			_ = summary_ws.cell(column = column_index, row = row_index, value = participant.stream)
			column_index += 1

			# Writes out the total correct for each of the categories
			for quiz_prefix in self.quiz_prefixes:
				category_total = participant.get_correct_count_for_category(quiz_prefix)
				_ = summary_ws.cell(column = column_index, row = row_index, value = int(category_total))
				column_index += 1

			# Writes out the average for each of the categories
			for scale_prefix in self.scale_prefixes:
				scale_average = participant.get_average_for_category_from_sum(scale_prefix)
				_ = summary_ws.cell(column = column_index, row = row_index, value = float(scale_average))
				column_index += 1

	def create_participant_worksheet(self, participant):
		""" Creates a worksheet that lists all of the given participants' responses """

		participant_title = "{} {}".format(participant.name, participant.stream)
		participant_ws = self.output_workbook.create_sheet(title = participant_title)

		heading_columns = ['Category', 'Question Number', 'Answer', 'Correct?']
		for (index, column) in enumerate(heading_columns):
			column_index = index + 1
			_ = participant_ws.cell(column = column_index, row = 1, value = column)

		for (index, response) in enumerate(participant.participant_responses):
			row_index = index + 2
			column_index = 1

			_ = participant_ws.cell(column = column_index, row = row_index, value = response.category)
			column_index += 1

			# Question number (different for the questions about integration (String, instead of a number)
			if response.category == 'Integration':
				_ = participant_ws.cell(column = column_index, row = row_index, value = response.question_number)
			else:
				_ = participant_ws.cell(column = column_index, row = row_index, value = int(response.question_number))
			column_index += 1

			# The response to the questions that are quizzes are mostly letters, the scale questions are numbers, integration are booleans
			
			answer_value = ''
			if response.is_quiz_question or response.category == 'Integration':
				answer_value = response.response_answer
			else:
				answer_value = int(response.response_answer)
			
			_ = participant_ws.cell(column = column_index, row = row_index, value = answer_value)
			column_index += 1

			# if the response being written to the cell is not a quiz question, an 'N/A' is displayed in the Correct? column
			correct_value = 'N/A'
			# If the response being written to the cell is a quiz question, we display whether or not it was correct
			if response.is_quiz_question:
				correct_value = 'Yes' if response.is_correct else 'No'

			_ = participant_ws.cell(column = column_index, row = row_index, value = correct_value)


	def perform_export(self, included_columns):
		""" Creates all of the worksheets for the participants and included columns. """

		self.included_columns = included_columns

		self.create_summary_worksheet()
		self.create_all_responses_worksheet()

		for participant in self.participants:
			self.create_participant_worksheet(participant)

		self.output_workbook.save(filename = 'output/responses.xlsx')









